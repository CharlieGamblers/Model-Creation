# predict_batch_excel.py  — batch prediction utilities aligned with model feature names (internal-ID only)

from __future__ import annotations

from pathlib import Path
import unicodedata
import pandas as pd
import xgboost as xgb
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .feature_engineering import solo_data_pull, load_average_bull_features
from .config import (
    RIDER_XLSX,
    BULL_XLSX,
    FINAL_DATA,
    MODEL_FILE,
    FEATURE_LIST,
    replacements,
    rider_replacements,
    DEFAULT_DATE,
    TEAM_COLORS,
    # model helpers
    load_model,
    model_feature_names,
)

# -------------------- Helpers --------------------

def clean_bull_name(stock: str, name: str, replacement_dict: dict) -> str:
    full = f"{str(stock).strip()} {str(name).strip()}"
    return replacement_dict.get(full.lower(), full)

def normalize(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(text).lower())
        if not unicodedata.combining(c)
    )

def _align_row_to_features(row_df: pd.DataFrame, features_expected: list[str]) -> pd.DataFrame:
    """
    Ensure row_df has exactly the columns the model expects:
      - add any missing as NaN
      - drop any extras
      - keep order identical to features_expected
      - cast to numeric for XGBoost
    """
    for col in features_expected:
        if col not in row_df.columns:
            row_df[col] = pd.NA
    row_df = row_df.loc[:, [c for c in features_expected if c in row_df.columns]]
    row_df = row_df.apply(pd.to_numeric, errors="coerce")
    return row_df

def _safe_lookup_dict(frame: pd.DataFrame, key_col: str, val_col: str) -> dict[str, int]:
    """
    Build a dict mapping normalized string(key_col) -> int(val_col), skipping NaNs.
    """
    if key_col not in frame.columns or val_col not in frame.columns:
        return {}
    out = {}
    for r, i in zip(frame[key_col], frame[val_col]):
        if pd.isna(r) or pd.isna(i):
            continue
        try:
            out[str(r).lower()] = int(i)
        except Exception:
            # skip rows with non-castable ids
            continue
    return out

# -------------------- Main: batch without Score --------------------

def predict_batch_from_excel(
    excel_path: Path,
    output_csv: Path,
    event_date: str = DEFAULT_DATE
) -> list[dict]:
    """
    Reads the Excel sheet, generates ride probabilities, and returns a list of dicts
    with 'game' (group of 10 rides), 'Team', 'Rider', 'Bull', 'Delivery', 'Outs', 'Probability'.
    Also writes a flat sheet to output_csv.

    INTERNAL-ID ONLY: Rider mapping uses rider_internal_id from rider_id_list.xlsx.
    Unknown riders are treated as 'new' (empty history) by passing rider_internal_id=None.
    """
    # Load lookups & data
    rid_map = pd.read_excel(RIDER_XLSX)
    bid_map = pd.read_excel(BULL_XLSX)[["bull", "bull_id"]]
    final_data = pd.read_csv(FINAL_DATA, parse_dates=["event_start_date"], low_memory=False)

    # Rider file must contain internal IDs
    # (Exported by build_dataset.py: columns include ['rider','rider_id','rider_internal_id','hand'])
    rider_name_col = "rider" if "rider" in rid_map.columns else rid_map.columns[0]
    if "rider_internal_id" not in rid_map.columns:
        raise KeyError(
            f"{RIDER_XLSX} must include 'rider_internal_id'. "
            "Re-run your data build/export step."
        )
    rid_lookup = _safe_lookup_dict(rid_map, rider_name_col, "rider_internal_id")
    bull_lookup = _safe_lookup_dict(bid_map, "bull", "bull_id")

    # Load model & expected feature names
    bst: xgb.Booster = load_model(MODEL_FILE)
    features_expected = model_feature_names(bst, FEATURE_LIST)

    # Known bull name replacements
    replacement_dict = dict(zip(
        replacements["Known_Discrepancies"].str.lower().str.strip(),
        replacements["Replacements"].str.strip()
    ))

    # Load worksheet (SectionWorksheet)
    df = pd.read_excel(excel_path, sheet_name="SectionWorksheet", skiprows=7)
    df = df.rename(columns=lambda x: str(x).strip())
    df = df.rename(columns={
        "Rider": "rider", "Bull": "bull", "Brand": "stock_no",
        "Del": "delivery", "Seq": "sequence", "Team": "team"
    })

    results: list[dict] = []
    failures: list[dict] = []
    start_date = pd.to_datetime(event_date)
    ride_counter = 0

    avg_bull_features = load_average_bull_features()

    for row in df.itertuples(index=False):
        raw_rider = str(getattr(row, "rider", "")).strip()
        raw_bull  = str(getattr(row, "bull",  "")).strip()
        stock_no  = str(getattr(row, "stock_no", "")).strip()

        if raw_rider == "" or raw_rider.lower() == "rider":
            continue
        if raw_bull == "" or raw_bull.lower() == "nan":
            continue
        if stock_no == "" or stock_no.lower() == "nan":
            continue

        # Group rides into sets of 10 → "game"
        ride_counter += 1
        game_id = (ride_counter - 1) // 10 + 1

        rider_key   = normalize(raw_rider)
        rider_clean = rider_replacements.get(rider_key, rider_key)  # normalized name key
        full_bull   = clean_bull_name(stock_no, raw_bull, replacement_dict).lower()

        rid_internal = rid_lookup.get(rider_clean, None)  # may be None (new rider)
        bid = bull_lookup.get(full_bull, None)

        prob = None
        outs_count = None
        reason = ""
        used_baseline = False

        try:
            if bid is None:
                print(f"[baseline] Bull not found: '{full_bull}'. Using average bull features.")
                used_baseline = True
                # Build rider features as normal, but use average bull features for bull columns
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=-1,  # Use dummy ID for missing bull
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )
                # Overwrite bull columns with average bull features
                for col in avg_bull_features.index:
                    if col in row_df.columns:
                        row_df[col] = avg_bull_features[col]
                outs_count = None
            else:
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=bid,
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )
                try:
                    outs_count = int(pd.to_numeric(final_data.loc[final_data["bull_id"] == bid, "out"]).max())
                except Exception:
                    outs_count = None

            # Align to model's exact feature set
            row_df = _align_row_to_features(row_df, features_expected)

            # Predict
            dm = xgb.DMatrix(row_df.values, feature_names=features_expected)
            prob = float(bst.predict(dm)[0])

        except Exception as e:
            failures.append({
                "game": game_id,
                "team": getattr(row, "team", ""),
                "rider": rider_clean,
                "bull": full_bull,
                "reason": str(e),
            })

        results.append({
            "game": game_id,
            "Team": getattr(row, "team", ""),
            "Rider": rider_clean.title(),
            "Bull": raw_bull,
            "Delivery": getattr(row, "delivery", ""),
            "Outs": outs_count,
            "Probability": round(prob * 100, 1) if prob is not None else None,
            "_UsedBaseline": used_baseline,  # internal only
        })

    # Save flat results
    df_out = pd.DataFrame(results)
    if "_UsedBaseline" in df_out.columns:
        df_out_excel = df_out.drop(columns=["_UsedBaseline"])
    else:
        df_out_excel = df_out
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df_out_excel.to_excel(output_csv, index=False)

    # Excel formatting
    wb = load_workbook(output_csv)
    ws = wb.active
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    team_col = headers.get("Team", "A")
    prob_col = headers.get("Probability", "F")
    team_colors = {team: TEAM_COLORS.get(team, "#FFFFFF") for team in df_out["Team"].unique()}
    missing_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")

    for i in range(2, ws.max_row + 1):
        team = ws[f"{team_col}{i}"].value
        prob_cell = ws[f"{prob_col}{i}"]
        hex_code = team_colors.get(team, "#FFFFFF")
        ws[f"{team_col}{i}"].fill = PatternFill(
            start_color="FF" + hex_code.lstrip("#"),
            end_color="FF" + hex_code.lstrip("#"),
            fill_type="solid"
        )
        # Highlight if baseline was used
        if df_out.iloc[i-2]["_UsedBaseline"]:
            prob_cell.fill = missing_fill
        prob_cell.number_format = '0.0"%"'

    # Optional AUS Monte Carlo sheet
    try:
        from .monte_carlo import add_aus_monte_carlo_sheet
        add_aus_monte_carlo_sheet(wb, df_out, team="AUS", sims=10000)
    except Exception:
        pass

    wb.save(output_csv)

    print(f"[✓] Saved {len(results)} predictions → {output_csv}")
    if failures:
        print(f"\n[!] Skipped {len(failures)} rows:")
        for f in failures:
            print(f" - Game {f['game']} | {f['team']} | {f['rider']} on {f['bull']} → {f['reason']}")
    return results

# -------------------- Main: batch with Score column --------------------

def predict_batch_from_excel_with_score(
    excel_path: Path,
    output_csv: Path,
    event_date: str = DEFAULT_DATE
) -> list[dict]:
    """
    Same as predict_batch_from_excel but also pulls the Score column (Excel column H
    if not named) and writes it to the output under 'Score'.
    """
    # Lookups & data
    rid_map = pd.read_excel(RIDER_XLSX)
    bid_map = pd.read_excel(BULL_XLSX)[["bull", "bull_id"]]
    final_data = pd.read_csv(FINAL_DATA, parse_dates=["event_start_date"], low_memory=False)

    # Rider internal-id mapping
    rider_name_col = "rider" if "rider" in rid_map.columns else rid_map.columns[0]
    if "rider_internal_id" not in rid_map.columns:
        raise KeyError(
            f"{RIDER_XLSX} must include 'rider_internal_id'. "
            "Re-run your data build/export step."
        )
    rid_lookup = _safe_lookup_dict(rid_map, rider_name_col, "rider_internal_id")
    bull_lookup = _safe_lookup_dict(bid_map, "bull", "bull_id")

    # Booster + feature names
    bst: xgb.Booster = load_model(MODEL_FILE)
    features_expected = model_feature_names(bst, FEATURE_LIST)

    replacement_dict = dict(zip(
        replacements["Known_Discrepancies"].str.lower().str.strip(),
        replacements["Replacements"].str.strip()
    ))

    # Input sheet
    df = pd.read_excel(excel_path, sheet_name="SectionWorksheet", skiprows=7)
    df = df.rename(columns=lambda x: str(x).strip())
    df = df.rename(columns={
        "Rider": "rider", "Bull": "bull", "Brand": "stock_no",
        "Del": "delivery", "Seq": "sequence", "Team": "team"
    })

    # Map 'Score' → 'score'
    if "Score" in df.columns:
        df = df.rename(columns={"Score": "score"})
    elif df.shape[1] >= 8:
        df = df.rename(columns={df.columns[7]: "score"})
    else:
        df["score"] = pd.NA

    results: list[dict] = []
    failures: list[dict] = []
    start_date = pd.to_datetime(event_date)
    ride_counter = 0

    avg_bull_features = load_average_bull_features()

    for row in df.itertuples(index=False):
        raw_rider = str(getattr(row, "rider", "")).strip()
        raw_bull  = str(getattr(row, "bull",  "")).strip()
        stock_no  = str(getattr(row, "stock_no", "")).strip()
        if raw_rider == "" or raw_rider.lower() == "rider" or raw_bull == "" or stock_no == "":
            continue

        ride_counter += 1
        game_id = (ride_counter - 1) // 10 + 1

        rider_key   = normalize(raw_rider)
        rider_clean = rider_replacements.get(rider_key, rider_key)
        full_bull   = clean_bull_name(stock_no, raw_bull, replacement_dict).lower()

        rid_internal = rid_lookup.get(rider_clean, None)
        bid = bull_lookup.get(full_bull, None)

        prob = None
        used_baseline = False
        try:
            if bid is None:
                print(f"[baseline] Bull not found: '{full_bull}'. Using average bull features.")
                used_baseline = True
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=-1,
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )
                for col in avg_bull_features.index:
                    if col in row_df.columns:
                        row_df[col] = avg_bull_features[col]
            else:
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=bid,
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )
            row_df = _align_row_to_features(row_df, features_expected)
            dm = xgb.DMatrix(row_df.values, feature_names=features_expected)
            prob = float(bst.predict(dm)[0])
        except Exception as e:
            failures.append({
                "game": game_id,
                "team": getattr(row, "team", ""),
                "rider": rider_clean,
                "bull": full_bull,
                "reason": str(e),
            })

        results.append({
            "game": game_id,
            "Team": getattr(row, "team", ""),
            "Rider": rider_clean.title(),
            "Bull": raw_bull,
            "Delivery": getattr(row, "delivery", ""),
            "Probability": round(prob * 100, 1) if prob is not None else None,
            "Score": getattr(row, "score", None),
            "_UsedBaseline": used_baseline,  # internal only
        })

    # Save output
    df_out = pd.DataFrame(results)
    if "_UsedBaseline" in df_out.columns:
        df_out_excel = df_out.drop(columns=["_UsedBaseline"])
    else:
        df_out_excel = df_out
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df_out_excel.to_excel(output_csv, index=False)

    # Formatting (Probability and Score columns)
    wb = load_workbook(output_csv)
    ws = wb.active
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    team_col = headers.get("Team", "A")
    prob_col = headers.get("Probability", "F")
    score_col = headers.get("Score", "G")
    team_colors = {team: TEAM_COLORS.get(team, "#FFFFFF") for team in df_out["Team"].unique()}
    missing_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    for i in range(2, ws.max_row + 1):
        team = ws[f"{team_col}{i}"].value
        prob_cell = ws[f"{prob_col}{i}"]
        score_cell = ws[f"{score_col}{i}"] if score_col in headers.values() else None
        hex_code = team_colors.get(team, "#FFFFFF")
        ws[f"{team_col}{i}"].fill = PatternFill(
            start_color="FF" + hex_code.lstrip("#"),
            end_color="FF" + hex_code.lstrip("#"),
            fill_type="solid"
        )
        # Highlight if baseline was used
        if df_out.iloc[i-2]["_UsedBaseline"]:
            prob_cell.fill = missing_fill
        prob_cell.number_format = '0.0"%"'
        if score_cell and score_cell.value not in (None, ""):
            score_cell.number_format = '0.0'

    wb.save(output_csv)

    print(f"[✓] Saved {len(results)} predictions (with Score) → {output_csv}")
    if failures:
        print(f"\n[!] Skipped {len(failures)} rows:")
        for f in failures:
            print(f" - Game {f['game']} | {f['team']} | {f['rider']} on {f['bull']} → {f['reason']}")
    return results
