from __future__ import annotations

from typing import Tuple, Dict
import numpy as np
import pandas as pd
from openpyxl.workbook.workbook import Workbook


def _extract_matchup(df_out: pd.DataFrame, team: str = "AUS") -> Tuple[list[float], list[float], str, str]:
    """From a flat predictions DataFrame (as written by batch_excel), pick the first
    game containing `team` and return probability vectors for both teams along with
    the opponent name.

    Returns (probs_team, probs_opp, team_name, opp_name) with probabilities in [0,1].
    Raises ValueError if the matchup cannot be determined.
    """
    if "game" not in df_out.columns:
        # infer game by blocks of 10 if missing
        df_out = df_out.copy()
        df_out["game"] = (np.arange(len(df_out)) // 10) + 1

    team_rows = df_out.loc[df_out["Team"] == team]
    if team_rows.empty:
        raise ValueError(f"Team '{team}' not found in predictions output")

    game_id = int(team_rows.iloc[0]["game"])  # first occurrence
    game_df = df_out.loc[df_out["game"] == game_id].copy()
    teams_in_game = [t for t in game_df["Team"].unique().tolist() if isinstance(t, str)]
    if len(teams_in_game) < 2:
        raise ValueError(f"Game {game_id} does not have two teams present")

    opp_name = next(t for t in teams_in_game if t != team)

    def _probs(d: pd.DataFrame) -> list[float]:
        # Probability column is in percent; fill missing with default 26.5
        p = pd.to_numeric(d["Probability"], errors="coerce").fillna(26.5) / 100.0
        # Keep top 5 rows if more present
        return p.iloc[:5].clip(0, 1).tolist()

    probs_team = _probs(game_df.loc[game_df["Team"] == team])
    probs_opp = _probs(game_df.loc[game_df["Team"] == opp_name])

    return probs_team, probs_opp, team, opp_name


def simulate_matchup(probs_team: list[float], probs_opp: list[float], sims: int = 10000, seed: int = 42) -> Dict:
    """Monte Carlo simulate a 5v5 game using per-ride success probabilities.
    Returns a dict with win/loss/draw counts and ride-count distributions per team.
    """
    n_a = len(probs_team)
    n_b = len(probs_opp)
    if n_a == 0 or n_b == 0:
        raise ValueError("Empty probability vectors for team or opponent")

    rng = np.random.default_rng(seed)

    wins_a = wins_b = draws = 0
    dist_a = np.zeros(n_a + 1, dtype=int)
    dist_b = np.zeros(n_b + 1, dtype=int)

    p_a = np.array(probs_team, dtype=float)
    p_b = np.array(probs_opp, dtype=float)

    for _ in range(sims):
        rides_a = (rng.random(n_a) < p_a).sum()
        rides_b = (rng.random(n_b) < p_b).sum()
        dist_a[rides_a] += 1
        dist_b[rides_b] += 1
        if rides_a > rides_b:
            wins_a += 1
        elif rides_b > rides_a:
            wins_b += 1
        else:
            draws += 1

    return {
        "wins_a": int(wins_a),
        "wins_b": int(wins_b),
        "draws": int(draws),
        "dist_a": dist_a.tolist(),
        "dist_b": dist_b.tolist(),
        "mean_a": float((np.arange(len(dist_a)) * dist_a).sum() / sims),
        "mean_b": float((np.arange(len(dist_b)) * dist_b).sum() / sims),
    }


def add_aus_monte_carlo_sheet(wb: Workbook, df_out: pd.DataFrame, team: str = "AUS", sims: int = 10000) -> None:
    """Append/replace a worksheet 'Monte Carlo' with AUS game simulation results."""
    title = "Monte Carlo"
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title)

    try:
        probs_team, probs_opp, team_name, opp_name = _extract_matchup(df_out, team)
        res = simulate_matchup(probs_team, probs_opp, sims=sims, seed=42)
    except ValueError as e:
        ws.cell(row=1, column=1).value = f"Error: {e}"
        return

    # Header
    ws.cell(row=1, column=1).value = f"{team_name} vs {opp_name} - {sims:,} simulations"
    ws.cell(row=2, column=1).value = "Outcome"
    ws.cell(row=2, column=2).value = "Count"
    ws.cell(row=2, column=3).value = "Percentage"

    # Results
    outcomes = [
        (f"{team_name} wins", res["wins_a"], res["wins_a"] / sims * 100),
        (f"{opp_name} wins", res["wins_b"], res["wins_b"] / sims * 100),
        ("Tie", res["draws"], res["draws"] / sims * 100),
    ]

    for i, (outcome, count, pct) in enumerate(outcomes, start=3):
        ws.cell(row=i, column=1).value = outcome
        ws.cell(row=i, column=2).value = count
        ws.cell(row=i, column=3).value = round(pct, 1)

    # Ride distributions
    row = 7
    ws.cell(row=row, column=1).value = "Ride Distributions"
    row += 1
    ws.cell(row=row, column=1).value = "Rides"
    ws.cell(row=row, column=2).value = f"{team_name} %"
    ws.cell(row=row, column=3).value = f"{opp_name} %"
    row += 1

    max_rides = max(len(res["dist_a"]), len(res["dist_b"]))
    for i in range(max_rides):
        team_pct = (res["dist_a"][i] / sims * 100) if i < len(res["dist_a"]) else 0
        opp_pct = (res["dist_b"][i] / sims * 100) if i < len(res["dist_b"]) else 0
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = round(team_pct, 1)
        ws.cell(row=row, column=3).value = round(opp_pct, 1)
        row += 1

    # Mean rides
    row += 1
    ws.cell(row=row, column=1).value = "Mean rides:"
    ws.cell(row=row, column=2).value = f"{team_name}: {res['mean_a']:.3f}"
    ws.cell(row=row, column=3).value = f"{opp_name}: {res['mean_b']:.3f}"


