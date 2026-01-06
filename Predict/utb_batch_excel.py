# predict_batch_excel.py  — batch prediction utilities aligned with model feature names (internal-ID only)

from __future__ import annotations

from pathlib import Path
import unicodedata
import pandas as pd
import xgboost as xgb
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .feature_engineering import solo_data_pull, load_average_bull_features
from .config import (
    RIDER_XLSX,
    BULL_XLSX,
    FINAL_DATA,
    MODEL_FILE,
    FEATURE_LIST,
    replacements,
    rider_replacements,
    DEFAULT_DATE,
    # model helpers
    load_model,
    model_feature_names,
)

# -------------------- Helpers --------------------


def clean_bull_name(stock: str, name: str, replacement_dict: dict) -> str:
    """
    Build a full bull key "Brand + Bull" and apply any known discrepancy replacements.
    """
    full = f"{str(stock).strip()} {str(name).strip()}"
    return replacement_dict.get(full.lower(), full)


def normalize(text: str) -> str:
    """
    Lowercase + strip accents so name matching is robust.
    """
    return "".join(
        c
        for c in unicodedata.normalize("NFKD", str(text).lower())
        if not unicodedata.combining(c)
    )


def _align_row_to_features(row_df: pd.DataFrame, features_expected: list[str]) -> pd.DataFrame:
    """
    Ensure row_df has exactly the columns the model expects:
      - add any missing as NaN
      - drop any extras
      - keep order identical to features_expected
      - cast to numeric for XGBoost
    """
    for col in features_expected:
        if col not in row_df.columns:
            row_df[col] = pd.NA

    row_df = row_df.loc[:, [c for c in features_expected if c in row_df.columns]]
    row_df = row_df.apply(pd.to_numeric, errors="coerce")
    return row_df


def _safe_lookup_dict(frame: pd.DataFrame, key_col: str, val_col: str) -> dict[str, int]:
    """
    Build a dict mapping normalized string(key_col) -> int(val_col), skipping NaNs.
    """
    if key_col not in frame.columns or val_col not in frame.columns:
        return {}
    out: dict[str, int] = {}
    for r, i in zip(frame[key_col], frame[val_col]):
        if pd.isna(r) or pd.isna(i):
            continue
        try:
            out[str(r).lower()] = int(i)
        except Exception:
            # skip rows with non-castable ids
            continue
    return out


# ============================================================
# 1. Main: batch without Score (Section / Seq / FullBull / Rider / Prob)
# ============================================================


def predict_batch_from_excel(
    excel_path: Path,
    output_csv: Path,
    event_date: str = DEFAULT_DATE,
) -> list[dict]:
    """
    Reads the SectionWorksheet sheet, generates ride probabilities, and returns a list of dicts
    with:
      - Section
      - Sequence
      - Bull (Brand + Bull)
      - Rider
      - Delivery
      - Probability

    INTERNAL-ID ONLY: Rider mapping uses rider_internal_id from rider_id_list.xlsx.
    Unknown riders are treated as 'new' (empty history) by passing rider_internal_id=None.
    """

    # ---------- Lookups & data ----------
    rid_map = pd.read_excel(RIDER_XLSX)
    bid_map = pd.read_excel(BULL_XLSX)[["bull", "bull_id"]]
    final_data = pd.read_csv(FINAL_DATA, parse_dates=["event_start_date"], low_memory=False)

    # Rider file must contain internal IDs
    # (Exported by build_dataset.py: columns include ['rider','rider_id','rider_internal_id','hand'])
    rider_name_col = "rider" if "rider" in rid_map.columns else rid_map.columns[0]
    if "rider_internal_id" not in rid_map.columns:
        raise KeyError(
            f"{RIDER_XLSX} must include 'rider_internal_id'. "
            "Re-run your data build/export step."
        )
    rid_lookup = _safe_lookup_dict(rid_map, rider_name_col, "rider_internal_id")
    bull_lookup = _safe_lookup_dict(bid_map, "bull", "bull_id")

    # Model + expected features
    bst: xgb.Booster = load_model(MODEL_FILE)
    features_expected = model_feature_names(bst, FEATURE_LIST)

    # Known bull name replacements (for "Brand + Bull")
    replacement_dict = dict(
        zip(
            replacements["Known_Discrepancies"].str.lower().str.strip(),
            replacements["Replacements"].str.strip(),
        )
    )

    # ---------- Input sheet (SectionWorksheet) ----------
    # header row is row 8 (0-index 7) in your file: columns like ['#','Rank','Bk#','Score','Team','Rider','Bull','TC','SC','Brand','Del','Sec','Seq']
    df = pd.read_excel(excel_path, sheet_name="SectionWorksheet", header=7)
    df = df.rename(columns=lambda x: str(x).strip())

    # Standardize key names we care about
    # (Team is intentionally ignored in this version)
    column_renames = {
        "Rider": "rider",
        "Bull": "bull",
        "Brand": "stock_no",
        "Del": "delivery",
        "Sec": "Sec",
        "Seq": "Seq",
    }
    df = df.rename(columns=column_renames)

    results: list[dict] = []
    failures: list[dict] = []
    start_date = pd.to_datetime(event_date)

    avg_bull_features = load_average_bull_features()

    # ---------- Row loop ----------
    for row in df.itertuples(index=False):
        raw_rider = str(getattr(row, "rider", "")).strip()
        raw_bull = str(getattr(row, "bull", "")).strip()
        stock_no = str(getattr(row, "stock_no", "")).strip()

        # Skip empty / header rows
        if raw_rider == "" or raw_rider.lower() == "rider":
            continue
        if raw_bull == "" or raw_bull.lower() in ("nan",):
            continue
        if stock_no == "" or stock_no.lower() in ("nan",):
            continue

        # Section + Sequence from worksheet
        section_val = getattr(row, "Sec", None)
        sequence_val = getattr(row, "Seq", None)
        delivery_val = getattr(row, "delivery", None)  # <<< added

        # Rider canonical key
        rider_key = normalize(raw_rider)
        rider_clean_key = rider_replacements.get(rider_key, rider_key)  # normalized key
        display_rider = str(raw_rider).strip()  # keep Excel's rider name for display

        # Full bull name (Brand + Bull) for both ID and display
        full_bull_clean = clean_bull_name(stock_no, raw_bull, replacement_dict).lower()
        full_bull_display = f"{stock_no} {raw_bull}".strip()

        # ID lookups
        rid_internal = rid_lookup.get(rider_clean_key, None)  # may be None (new rider)
        bid = bull_lookup.get(full_bull_clean, None)

        prob = None
        used_baseline = False

        try:
            if bid is None:
                # Unknown bull → use average bull features
                print(f"[baseline] Bull not found: '{full_bull_clean}'. Using average bull features.")
                used_baseline = True

                # Build rider features with dummy bull_id, then overwrite bull columns with averages
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=-1,  # dummy for missing bull
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )
                for col in avg_bull_features.index:
                    if col in row_df.columns:
                        row_df[col] = avg_bull_features[col]
            else:
                # Normal rider+bull pull
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=bid,
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )

            # Align to model's exact feature set
            row_df = _align_row_to_features(row_df, features_expected)

            # Predict
            dm = xgb.DMatrix(row_df.values, feature_names=features_expected)
            prob = float(bst.predict(dm)[0])

        except Exception as e:
            failures.append(
                {
                    "section": section_val,
                    "sequence": sequence_val,
                    "rider": rider_clean_key,
                    "bull": full_bull_clean,
                    "reason": str(e),
                }
            )

        results.append(
            {
                "Section": section_val,
                "Sequence": sequence_val,
                "Bull": full_bull_display,
                "Rider": display_rider,
                "Delivery": delivery_val,  # <<< added
                "Probability": round(prob * 100, 1) if prob is not None else None,
                "_UsedBaseline": used_baseline,  # internal only (for formatting)
            }
        )

    # ---------- Save flat results ----------
    df_out = pd.DataFrame(results)
    if "_UsedBaseline" in df_out.columns:
        df_out_excel = df_out.drop(columns=["_UsedBaseline"])
    else:
        df_out_excel = df_out

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df_out_excel.to_excel(output_csv, index=False)

    # ---------- Excel formatting ----------
    wb = load_workbook(output_csv)
    ws = wb.active
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    prob_col = headers.get("Probability")

    # Light red fill for baseline bulls (missing in lookup)
    missing_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")

    if prob_col is not None:
        for i in range(2, ws.max_row + 1):
            prob_cell = ws[f"{prob_col}{i}"]

            # Highlight baseline rows
            if "_UsedBaseline" in df_out.columns and df_out.iloc[i - 2]["_UsedBaseline"]:
                prob_cell.fill = missing_fill

            # Format as percentage
            prob_cell.number_format = '0.0"%"'

    wb.save(output_csv)

    print(f"[✓] Saved {len(results)} predictions → {output_csv}")
    if failures:
        print(f"\n[!] Skipped {len(failures)} rows:")
        for f in failures:
            print(
                f" - Sec {f['section']} | Seq {f['sequence']} | "
                f"{f['rider']} on {f['bull']} → {f['reason']}"
            )
    return results



# ============================================================
# 2. Main: batch WITH Score (adds Score column)
# ============================================================


def predict_batch_from_excel_with_score(
    excel_path: Path,
    output_csv: Path,
    event_date: str = DEFAULT_DATE,
) -> list[dict]:
    """
    Same as predict_batch_from_excel but also pulls the 'Score' column
    from the SectionWorksheet and writes it under 'Score'.
    Output columns:
      - Section
      - Sequence
      - Bull (Brand + Bull)
      - Rider
      - Probability
      - Score
    """

    # ---------- Lookups & data ----------
    rid_map = pd.read_excel(RIDER_XLSX)
    bid_map = pd.read_excel(BULL_XLSX)[["bull", "bull_id"]]
    final_data = pd.read_csv(FINAL_DATA, parse_dates=["event_start_date"], low_memory=False)

    rider_name_col = "rider" if "rider" in rid_map.columns else rid_map.columns[0]
    if "rider_internal_id" not in rid_map.columns:
        raise KeyError(
            f"{RIDER_XLSX} must include 'rider_internal_id'. "
            "Re-run your data build/export step."
        )
    rid_lookup = _safe_lookup_dict(rid_map, rider_name_col, "rider_internal_id")
    bull_lookup = _safe_lookup_dict(bid_map, "bull", "bull_id")

    bst: xgb.Booster = load_model(MODEL_FILE)
    features_expected = model_feature_names(bst, FEATURE_LIST)

    replacement_dict = dict(
        zip(
            replacements["Known_Discrepancies"].str.lower().str.strip(),
            replacements["Replacements"].str.strip(),
        )
    )

    # ---------- Input sheet ----------
    df = pd.read_excel(excel_path, sheet_name="SectionWorksheet", header=7)
    df = df.rename(columns=lambda x: str(x).strip())

    column_renames = {
        "Rider": "rider",
        "Bull": "bull",
        "Brand": "stock_no",
        "Del": "delivery",
        "Sec": "Sec",
        "Seq": "Seq",
    }
    df = df.rename(columns=column_renames)

    # Map 'Score' → 'score' (if there is a Score column)
    if "Score" in df.columns:
        df = df.rename(columns={"Score": "score"})
    else:
        # If score column name is unknown but exists at usual position, use that.
        # You can tweak this if your layout changes.
        if df.shape[1] >= 8:
            df = df.rename(columns={df.columns[7]: "score"})
        else:
            df["score"] = pd.NA

    results: list[dict] = []
    failures: list[dict] = []
    start_date = pd.to_datetime(event_date)

    avg_bull_features = load_average_bull_features()

    # ---------- Row loop ----------
    for row in df.itertuples(index=False):
        raw_rider = str(getattr(row, "rider", "")).strip()
        raw_bull = str(getattr(row, "bull", "")).strip()
        stock_no = str(getattr(row, "stock_no", "")).strip()

        if raw_rider == "" or raw_rider.lower() == "rider":
            continue
        if raw_bull == "" or raw_bull.lower() in ("nan",):
            continue
        if stock_no == "" or stock_no.lower() in ("nan",):
            continue

        section_val = getattr(row, "Sec", None)
        sequence_val = getattr(row, "Seq", None)
        score_val = getattr(row, "score", None)

        rider_key = normalize(raw_rider)
        rider_clean_key = rider_replacements.get(rider_key, rider_key)
        display_rider = str(raw_rider).strip()

        full_bull_clean = clean_bull_name(stock_no, raw_bull, replacement_dict).lower()
        full_bull_display = f"{stock_no} {raw_bull}".strip()

        rid_internal = rid_lookup.get(rider_clean_key, None)
        bid = bull_lookup.get(full_bull_clean, None)

        prob = None
        used_baseline = False

        try:
            if bid is None:
                print(f"[baseline] Bull not found: '{full_bull_clean}'. Using average bull features.")
                used_baseline = True
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=-1,
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )
                for col in avg_bull_features.index:
                    if col in row_df.columns:
                        row_df[col] = avg_bull_features[col]
            else:
                row_df = solo_data_pull(
                    final_data,
                    rider_id=None,
                    bull_id=bid,
                    rider_internal_id=rid_internal,
                    start_date=start_date,
                )

            row_df = _align_row_to_features(row_df, features_expected)
            dm = xgb.DMatrix(row_df.values, feature_names=features_expected)
            prob = float(bst.predict(dm)[0])
        except Exception as e:
            failures.append(
                {
                    "section": section_val,
                    "sequence": sequence_val,
                    "rider": rider_clean_key,
                    "bull": full_bull_clean,
                    "reason": str(e),
                }
            )

        results.append(
            {
                "Section": section_val,
                "Sequence": sequence_val,
                "Bull": full_bull_display,
                "Rider": display_rider,
                "Probability": round(prob * 100, 1) if prob is not None else None,
                "Score": score_val,
                "_UsedBaseline": used_baseline,
            }
        )

    # ---------- Save results ----------
    df_out = pd.DataFrame(results)
    if "_UsedBaseline" in df_out.columns:
        df_out_excel = df_out.drop(columns=["_UsedBaseline"])
    else:
        df_out_excel = df_out

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df_out_excel.to_excel(output_csv, index=False)

    # ---------- Formatting ----------
    wb = load_workbook(output_csv)
    ws = wb.active
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    prob_col = headers.get("Probability")
    score_col = headers.get("Score")

    missing_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")

    for i in range(2, ws.max_row + 1):
        # Probability column formatting
        if prob_col is not None:
            prob_cell = ws[f"{prob_col}{i}"]
            if "_UsedBaseline" in df_out.columns and df_out.iloc[i - 2]["_UsedBaseline"]:
                prob_cell.fill = missing_fill
            prob_cell.number_format = '0.0"%"'

        # Score column formatting (if present)
        if score_col is not None:
            score_cell = ws[f"{score_col}{i}"]
            if score_cell.value not in (None, ""):
                score_cell.number_format = "0.0"

    wb.save(output_csv)

    print(f"[✓] Saved {len(results)} predictions (with Score) → {output_csv}")
    if failures:
        print(f"\n[!] Skipped {len(failures)} rows:")
        for f in failures:
            print(
                f" - Sec {f['section']} | Seq {f['sequence']} | "
                f"{f['rider']} on {f['bull']} → {f['reason']}"
            )
    return results
